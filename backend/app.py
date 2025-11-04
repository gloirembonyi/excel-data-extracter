from fastapi import FastAPI, File, UploadFile, HTTPException, BackgroundTasks
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
import pandas as pd
import numpy as np
import io
from typing import List, Dict, Any, Optional
import json
import os
import tempfile
from datetime import datetime
import logging
import asyncio
import aiohttp
import base64
from concurrent.futures import ThreadPoolExecutor, as_completed
import time
from dataclasses import dataclass
from enum import Enum
# import asyncpg
from sqlalchemy import create_engine, Column, String, Integer, DateTime, Text, Float
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker
from sqlalchemy.dialects.postgresql import UUID
import uuid

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Enums and Data Classes
class ProcessingStatus(Enum):
    PENDING = "pending"
    PROCESSING = "processing"
    COMPLETED = "completed"
    FAILED = "failed"
    RETRYING = "retrying"

@dataclass
class ImageProcessingResult:
    id: str
    filename: str
    status: ProcessingStatus
    extracted_text: str = ""
    table_data: List[Dict[str, Any]] = None
    error_message: str = ""
    processing_time: float = 0.0
    retry_count: int = 0
    api_key_used: str = ""

@dataclass
class BatchProcessingJob:
    id: str
    total_images: int
    completed_images: int
    failed_images: int
    status: ProcessingStatus
    results: List[ImageProcessingResult] = None
    created_at: datetime = None
    completed_at: datetime = None

# Database setup
DATABASE_URL = os.getenv("DATABASE_URL", "postgresql://neondb_owner:npg_zXrusVC6D3NZ@ep-curly-voice-ads6ydqv-pooler.c-2.us-east-1.aws.neon.tech/neondb?sslmode=require&channel_binding=require")
engine = create_engine(DATABASE_URL, pool_pre_ping=True, pool_recycle=300, pool_size=10, max_overflow=20)
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
Base = declarative_base()

# Test database connection
try:
    with engine.connect() as conn:
        conn.execute("SELECT 1")
    logger.info("Database connection successful")
except Exception as e:
    logger.error(f"Database connection failed: {e}")

# Database Models
class Project(Base):
    __tablename__ = "projects"
    
    id = Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    name = Column(String, nullable=False)
    description = Column(Text)
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

class MasterDataItem(Base):
    __tablename__ = "master_data_items"
    
    id = Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    project_id = Column(UUID(as_uuid=True), nullable=False)
    item_description = Column(String, nullable=False)
    serial_number = Column(String, nullable=False)
    tag_number = Column(String, nullable=False)
    quantity = Column(Integer, default=1)
    status = Column(String, default="New")
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

class ProcessedImage(Base):
    __tablename__ = "processed_images"
    
    id = Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    project_id = Column(UUID(as_uuid=True), nullable=False)
    filename = Column(String, nullable=False)
    extracted_data = Column(Text)  # JSON string
    matched_data = Column(Text)    # JSON string
    processing_status = Column(String, default="pending")
    created_at = Column(DateTime, default=datetime.utcnow)

class Dataset(Base):
    __tablename__ = "datasets"
    
    id = Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    name = Column(String, nullable=False)
    description = Column(Text)
    file_count = Column(Integer, default=0)
    total_rows = Column(Integer, default=0)
    files = Column(Text)  # JSON string of file names
    data = Column(Text)   # JSON string of combined data
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

# Create tables
Base.metadata.create_all(bind=engine)

# Global storage for batch jobs (in production, use Redis or database)
batch_jobs: Dict[str, BatchProcessingJob] = {}

app = FastAPI(title="Excel Table Maker API", version="1.0.0")

# CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000", "http://127.0.0.1:3000"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.get("/")
async def root():
    return {"message": "Excel Table Maker API is running"}

@app.post("/upload-excel")
async def upload_excel(file: UploadFile = File(...)):
    """Upload and parse Excel/CSV file"""
    try:
        # Read file content
        content = await file.read()
        
        # Determine file type and parse accordingly
        if file.filename.endswith('.csv'):
            # Parse CSV
            df = pd.read_csv(io.BytesIO(content))
        elif file.filename.endswith(('.xlsx', '.xls')):
            # Parse Excel
            df = pd.read_excel(io.BytesIO(content))
        else:
            raise HTTPException(status_code=400, detail="Unsupported file type. Please upload CSV or Excel files.")
        
        # Convert to list of dictionaries
        data = df.to_dict('records')
        columns = list(df.columns)
        
        return {
            "success": True,
            "message": f"Successfully processed {file.filename}",
            "data": data,
            "columns": columns,
            "rows": len(data),
            "filename": file.filename
        }
    except Exception as e:
        logger.error(f"Error processing file: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")

@app.post("/compare-data")
async def compare_data(
    data1: List[Dict[str, Any]],
    data2: List[Dict[str, Any]],
    key_field: Optional[str] = None
):
    """Compare two datasets"""
    try:
        df1 = pd.DataFrame(data1)
        df2 = pd.DataFrame(data2)
        
        if key_field and key_field in df1.columns and key_field in df2.columns:
            # Merge on key field
            merged = pd.merge(df1, df2, on=key_field, how='outer', indicator=True)
            
            common = merged[merged['_merge'] == 'both'].drop('_merge', axis=1).to_dict('records')
            only_in_first = merged[merged['_merge'] == 'left_only'].drop('_merge', axis=1).to_dict('records')
            only_in_second = merged[merged['_merge'] == 'right_only'].drop('_merge', axis=1).to_dict('records')
            
            # Find differences in common records
            differences = []
            for _, row in merged[merged['_merge'] == 'both'].iterrows():
                for col in df1.columns:
                    if col in df2.columns and col != key_field:
                        if row[f'{col}_x'] != row[f'{col}_y']:
                            differences.append({
                                "key": row[key_field],
                                "field": col,
                                "value1": row[f'{col}_x'],
                                "value2": row[f'{col}_y']
                            })
        else:
            # Simple comparison without key field
            common = []
            only_in_first = data1
            only_in_second = data2
            differences = []
        
        return {
            "success": True,
            "message": "Data comparison completed",
            "common": common,
            "only_in_first": only_in_first,
            "only_in_second": only_in_second,
            "differences": differences,
            "summary": {
                "total_1": len(data1),
                "total_2": len(data2),
                "common_count": len(common),
                "only_in_first_count": len(only_in_first),
                "only_in_second_count": len(only_in_second),
                "differences_count": len(differences)
            }
        }
    except Exception as e:
        logger.error(f"Error comparing data: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error comparing data: {str(e)}")

@app.post("/merge-data")
async def merge_data(
    data1: List[Dict[str, Any]],
    data2: List[Dict[str, Any]],
    key_field: Optional[str] = None,
    merge_type: str = "outer"  # outer, inner, left, right
):
    """Merge two datasets"""
    try:
        df1 = pd.DataFrame(data1)
        df2 = pd.DataFrame(data2)
        
        if key_field and key_field in df1.columns and key_field in df2.columns:
            # Merge on key field
            if merge_type == "inner":
                merged = pd.merge(df1, df2, on=key_field, how='inner')
            elif merge_type == "left":
                merged = pd.merge(df1, df2, on=key_field, how='left')
            elif merge_type == "right":
                merged = pd.merge(df1, df2, on=key_field, how='right')
            else:  # outer
                merged = pd.merge(df1, df2, on=key_field, how='outer')
        else:
            # Simple concatenation
            merged = pd.concat([df1, df2], ignore_index=True)
        
        data = merged.to_dict('records')
        columns = list(merged.columns)
        
        return {
            "success": True,
            "data": data,
            "columns": columns,
            "rows": len(data)
        }
    except Exception as e:
        logger.error(f"Error merging data: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error merging data: {str(e)}")

@app.post("/filter-data")
async def filter_data(
    data: List[Dict[str, Any]],
    filters: Dict[str, Any]
):
    """Filter data based on criteria"""
    try:
        df = pd.DataFrame(data)
        
        # Apply filters
        for field, value in filters.items():
            if value is not None and value != '' and field in df.columns:
                if isinstance(value, str):
                    df = df[df[field].astype(str).str.lower().str.contains(str(value).lower(), na=False)]
                else:
                    df = df[df[field] == value]
        
        filtered_data = df.to_dict('records')
        
        return {
            "success": True,
            "data": filtered_data,
            "original_rows": len(data),
            "filtered_rows": len(filtered_data)
        }
    except Exception as e:
        logger.error(f"Error filtering data: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error filtering data: {str(e)}")

@app.post("/sort-data")
async def sort_data(
    data: List[Dict[str, Any]],
    sort_field: str,
    ascending: bool = True
):
    """Sort data by field"""
    try:
        df = pd.DataFrame(data)
        
        if sort_field not in df.columns:
            raise HTTPException(status_code=400, detail=f"Field '{sort_field}' not found in data")
        
        sorted_df = df.sort_values(by=sort_field, ascending=ascending)
        sorted_data = sorted_df.to_dict('records')
        
        return {
            "success": True,
            "data": sorted_data
        }
    except Exception as e:
        logger.error(f"Error sorting data: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error sorting data: {str(e)}")

@app.post("/export-excel")
async def export_excel(data: List[Dict[str, Any]], filename: str = "exported_data.xlsx"):
    """Export data to Excel file"""
    try:
        if not data:
            raise HTTPException(status_code=400, detail="No data to export")
        
        df = pd.DataFrame(data)
        
        # Create Excel file in memory
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Data', index=False)
        
        output.seek(0)
        
        return StreamingResponse(
            io.BytesIO(output.read()),
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        logger.error(f"Error exporting Excel: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error exporting Excel: {str(e)}")

@app.post("/export-csv")
async def export_csv(data: List[Dict[str, Any]], filename: str = "exported_data.csv"):
    """Export data to CSV file"""
    try:
        if not data:
            raise HTTPException(status_code=400, detail="No data to export")
        
        df = pd.DataFrame(data)
        
        # Create CSV content
        csv_content = df.to_csv(index=False)
        
        # Create temporary file
        with tempfile.NamedTemporaryFile(mode='w', delete=False, suffix='.csv', encoding='utf-8') as tmp_file:
            tmp_file.write(csv_content)
            
            return FileResponse(
                tmp_file.name,
                media_type='text/csv',
                filename=filename
            )
    except Exception as e:
        logger.error(f"Error exporting CSV: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error exporting CSV: {str(e)}")

# AI Processing Functions
async def extract_text_from_image(image_data: bytes, api_key: str) -> str:
    """Extract text from image using Gemini API"""
    try:
        # Encode image to base64
        image_base64 = base64.b64encode(image_data).decode('utf-8')
        
        async with aiohttp.ClientSession() as session:
            url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-1.5-flash:generateContent?key={api_key}"
            
            payload = {
                "contents": [{
                    "parts": [{
                        "text": "Extract all text from this image, including any tables or structured data. Return the raw text exactly as it appears."
                    }, {
                        "inline_data": {
                            "mime_type": "image/jpeg",
                            "data": image_base64
                        }
                    }]
                }],
                "generationConfig": {
                    "temperature": 0.1,
                    "topK": 32,
                    "topP": 1,
                    "maxOutputTokens": 4096,
                }
            }
            
            async with session.post(url, json=payload) as response:
                if response.status == 200:
                    result = await response.json()
                    return result['candidates'][0]['content']['parts'][0]['text']
                else:
                    error_text = await response.text()
                    raise Exception(f"API Error: {response.status} - {error_text}")
                    
    except Exception as e:
        logger.error(f"Error extracting text: {str(e)}")
        raise e

async def format_text_as_table(text: str, api_key: str) -> List[Dict[str, Any]]:
    """Format extracted text as structured table data"""
    try:
        prompt = f"""Extract equipment/inventory data from this text and format it as a JSON array. Each item should have these exact columns:

Required columns:
- Item_Description: The name/description of the equipment (e.g., "Screen", "CPU", "hp screen")
- Quantity: Number of items (extract numbers only, e.g., "1", "2")
- Serial_Number: Serial number or identifier (e.g., "1H35070V93", "1HF5MOW3X", "1HFSAMOW2R", "AH 35070 V2H")
- Tag_Number: Tag number or code (e.g., "MOHDIG125/SCR587", "MOH/AIG125/CPU 1131", "CPU2024", "SCR 513")
- Status: Status of the item (e.g., "New", "CPU 1127", "2014")

IMPORTANT PATTERNS TO RECOGNIZE:

1. CPU Items:
   - "1 CPU 1 1HF5MOW3X MOH/ΔIG1951 CPU 1127"
   - "1 CPU 1 1HF5110YJZ MOH/AIG125/CPU 1131"
   - "1 CPU 1 1HFSAMOW2R CPU2024"

2. Screen Items:
   - "1 Screen 1 1H35070V93 MOHDIG125/SCR587"
   - "1 Screen 1 1H35070V93 SCR 513"
   - "1 Screen 1H 350 70T 94 MOHIDIG125/SCR234"

3. Serial Number Patterns:
   - Starts with "1H", "AH", "1HF", "1HFS" followed by alphanumeric characters
   - May contain spaces: "1H 350 70T 94", "AH 35070 V2H"
   - Examples: "1H35070V93", "1HF5MOW3X", "1HFSAMOW2R"

4. Tag Number Patterns:
   - MOH patterns: "MOHDIG125/SCR587", "MOH/AIG125/CPU 1131", "MOH/ΔIG1951"
   - CPU patterns: "CPU2024", "CPU109P", "CPU 1131"
   - SCR patterns: "SCR 513", "SCR587", "SCR011"

Instructions:
1. Look for the specific patterns above
2. Extract each line as a separate row
3. Parse Serial_Number and Tag_Number accurately - they are distinct fields
4. Handle spaces in serial numbers correctly
5. If any column is missing, use empty string ""
6. Return ONLY a valid JSON array

Text to process:
{text}

Example format:
[
  {{
    "Item_Description": "CPU",
    "Quantity": "1",
    "Serial_Number": "1HF5MOW3X",
    "Tag_Number": "MOH/ΔIG1951",
    "Status": "CPU 1127"
  }},
  {{
    "Item_Description": "Screen",
    "Quantity": "1",
    "Serial_Number": "1H35070V93",
    "Tag_Number": "MOHDIG125/SCR587",
    "Status": "New"
  }}
]

Return only the JSON array, no additional text."""

        async with aiohttp.ClientSession() as session:
            url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-1.5-flash:generateContent?key={api_key}"
            
            payload = {
                "contents": [{
                    "parts": [{"text": prompt}]
                }],
                "generationConfig": {
                    "temperature": 0.1,
                    "topK": 32,
                    "topP": 1,
                    "maxOutputTokens": 4096,
                }
            }
            
            async with session.post(url, json=payload) as response:
                if response.status == 200:
                    result = await response.json()
                    json_text = result['candidates'][0]['content']['parts'][0]['text']
                    
                    # Clean and parse JSON
                    json_text = json_text.strip()
                    if json_text.startswith('```json'):
                        json_text = json_text[7:]
                    if json_text.endswith('```'):
                        json_text = json_text[:-3]
                    
                    return json.loads(json_text)
                else:
                    error_text = await response.text()
                    raise Exception(f"API Error: {response.status} - {error_text}")
                    
    except Exception as e:
        logger.error(f"Error formatting text: {str(e)}")
        # Use fallback parsing
        return fallback_parse_equipment_data(text)

def fallback_parse_equipment_data(text: str) -> List[Dict[str, Any]]:
    """Fallback parsing when AI fails"""
    lines = text.split('\n')
    result = []
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
            
        # Simple pattern matching for equipment data
        # Look for patterns like "1 Screen 1 1H35070V93 MOHDIG125/SCR587 New"
        parts = line.split()
        if len(parts) >= 4:
            try:
                # Try to extract quantity
                quantity = "1"
                if parts[0].isdigit():
                    quantity = parts[0]
                    parts = parts[1:]
                
                # Extract item description
                item_desc = parts[0] if parts else ""
                
                # Look for serial number (starts with 1H, AH, etc.)
                serial_number = ""
                tag_number = ""
                status = "New"
                
                for i, part in enumerate(parts[1:], 1):
                    if part.startswith(('1H', 'AH', '1HF')):
                        serial_number = part
                        # Get tag number (next part or parts)
                        if i + 1 < len(parts):
                            tag_number = parts[i + 1]
                        if i + 2 < len(parts):
                            status = parts[i + 2]
                        break
                
                if item_desc and serial_number:
                    result.append({
                        "Item_Description": item_desc,
                        "Quantity": quantity,
                        "Serial_Number": serial_number,
                        "Tag_Number": tag_number,
                        "Status": status
                    })
            except Exception as e:
                logger.warning(f"Error parsing line '{line}': {e}")
                continue
    
    return result

# API Keys (in production, use environment variables)
API_KEYS = [
    "AIzaSyBDqoYqf5QS9qm8Z99rOZsRZi2ChA_Dw8w",
    "AIzaSyBDqoYqf5QS9qm8Z99rOZsRZi2ChA_Dw8w",  # Duplicate for load balancing
    "AIzaSyBDqoYqf5QS9qm8Z99rOZsRZi2ChA_Dw8w",
    "AIzaSyBDqoYqf5QS9qm8Z99rOZsRZi2ChA_Dw8w",
    "AIzaSyBDqoYqf5QS9qm8Z99rOZsRZi2ChA_Dw8w",
    "AIzaSyBDqoYqf5QS9qm8Z99rOZsRZi2ChA_Dw8w"
]

def get_next_api_key() -> str:
    """Get next available API key (round-robin)"""
    return API_KEYS[0]  # For now, use the same key

async def process_single_image(image_data: bytes, filename: str, image_id: str, max_retries: int = 3) -> ImageProcessingResult:
    """Process a single image with retry logic"""
    result = ImageProcessingResult(
        id=image_id,
        filename=filename,
        status=ProcessingStatus.PROCESSING
    )
    
    start_time = time.time()
    
    for attempt in range(max_retries):
        try:
            api_key = get_next_api_key()
            result.api_key_used = api_key
            
            # Extract text
            extracted_text = await extract_text_from_image(image_data, api_key)
            result.extracted_text = extracted_text
            
            # Format as table
            table_data = await format_text_as_table(extracted_text, api_key)
            result.table_data = table_data
            
            result.status = ProcessingStatus.COMPLETED
            result.processing_time = time.time() - start_time
            logger.info(f"Successfully processed {filename} in {result.processing_time:.2f}s")
            return result
            
        except Exception as e:
            result.retry_count = attempt + 1
            result.error_message = str(e)
            logger.warning(f"Attempt {attempt + 1} failed for {filename}: {str(e)}")
            
            if attempt < max_retries - 1:
                result.status = ProcessingStatus.RETRYING
                # Exponential backoff
                await asyncio.sleep(2 ** attempt)
            else:
                result.status = ProcessingStatus.FAILED
                logger.error(f"All attempts failed for {filename}: {str(e)}")
    
    result.processing_time = time.time() - start_time
    return result

async def process_batch_images(images: List[UploadFile], job_id: str, max_concurrent: int = 5):
    """Process multiple images in parallel with controlled concurrency"""
    job = batch_jobs[job_id]
    job.status = ProcessingStatus.PROCESSING
    job.results = []
    
    # Create semaphore to limit concurrent processing
    semaphore = asyncio.Semaphore(max_concurrent)
    
    async def process_with_semaphore(image_data: bytes, filename: str, image_id: str):
        async with semaphore:
            return await process_single_image(image_data, filename, image_id)
    
    # Process images in parallel
    tasks = []
    for i, image in enumerate(images):
        image_data = await image.read()
        image_id = f"{job_id}_image_{i}"
        task = process_with_semaphore(image_data, image.filename, image_id)
        tasks.append(task)
    
    # Wait for all tasks to complete
    results = await asyncio.gather(*tasks, return_exceptions=True)
    
    # Process results
    for i, result in enumerate(results):
        if isinstance(result, Exception):
            # Create failed result for exceptions
            failed_result = ImageProcessingResult(
                id=f"{job_id}_image_{i}",
                filename=images[i].filename,
                status=ProcessingStatus.FAILED,
                error_message=str(result)
            )
            job.results.append(failed_result)
            job.failed_images += 1
        else:
            job.results.append(result)
            if result.status == ProcessingStatus.COMPLETED:
                job.completed_images += 1
            else:
                job.failed_images += 1
    
    # Update job status
    if job.failed_images == 0:
        job.status = ProcessingStatus.COMPLETED
    elif job.completed_images > 0:
        job.status = ProcessingStatus.COMPLETED  # Partial success
    else:
        job.status = ProcessingStatus.FAILED
    
    job.completed_at = datetime.now()
    logger.info(f"Batch job {job_id} completed: {job.completed_images} successful, {job.failed_images} failed")

# Batch Processing Endpoints
@app.post("/batch-process-images")
async def batch_process_images(
    background_tasks: BackgroundTasks,
    images: List[UploadFile] = File(...),
    max_concurrent: int = 5
):
    """Start batch processing of multiple images"""
    try:
        if len(images) > 100:
            raise HTTPException(status_code=400, detail="Maximum 100 images allowed per batch")
        
        # Create batch job
        job_id = f"batch_{int(time.time())}"
        job = BatchProcessingJob(
            id=job_id,
            total_images=len(images),
            completed_images=0,
            failed_images=0,
            status=ProcessingStatus.PENDING,
            created_at=datetime.now()
        )
        batch_jobs[job_id] = job
        
        # Start background processing
        background_tasks.add_task(process_batch_images, images, job_id, max_concurrent)
        
        return {
            "success": True,
            "job_id": job_id,
            "total_images": len(images),
            "status": "processing_started"
        }
        
    except Exception as e:
        logger.error(f"Error starting batch processing: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error starting batch processing: {str(e)}")

@app.get("/batch-status/{job_id}")
async def get_batch_status(job_id: str):
    """Get status of a batch processing job"""
    if job_id not in batch_jobs:
        raise HTTPException(status_code=404, detail="Job not found")
    
    job = batch_jobs[job_id]
    return {
        "success": True,
        "job_id": job_id,
        "status": job.status.value,
        "total_images": job.total_images,
        "completed_images": job.completed_images,
        "failed_images": job.failed_images,
        "progress_percentage": (job.completed_images + job.failed_images) / job.total_images * 100,
        "created_at": job.created_at.isoformat() if job.created_at else None,
        "completed_at": job.completed_at.isoformat() if job.completed_at else None
    }

@app.get("/batch-results/{job_id}")
async def get_batch_results(job_id: str):
    """Get results of a completed batch processing job"""
    if job_id not in batch_jobs:
        raise HTTPException(status_code=404, detail="Job not found")
    
    job = batch_jobs[job_id]
    
    if job.status not in [ProcessingStatus.COMPLETED, ProcessingStatus.FAILED]:
        raise HTTPException(status_code=400, detail="Job not completed yet")
    
    return {
        "success": True,
        "job_id": job_id,
        "status": job.status.value,
        "total_images": job.total_images,
        "completed_images": job.completed_images,
        "failed_images": job.failed_images,
        "results": [
            {
                "id": result.id,
                "filename": result.filename,
                "status": result.status.value,
                "extracted_text": result.extracted_text,
                "table_data": result.table_data,
                "error_message": result.error_message,
                "processing_time": result.processing_time,
                "retry_count": result.retry_count,
                "api_key_used": result.api_key_used
            }
            for result in job.results
        ]
    }

# Database endpoints
@app.post("/projects")
async def create_project(request: Dict[str, Any]):
    """Create a new project"""
    try:
        name = request.get("name")
        description = request.get("description")
        
        if not name:
            raise HTTPException(status_code=400, detail="Project name is required")
        
        db = SessionLocal()
        project = Project(name=name, description=description)
        db.add(project)
        db.commit()
        db.refresh(project)
        db.close()
        
        return {
            "id": str(project.id),
            "name": project.name,
            "description": project.description,
            "created_at": project.created_at.isoformat(),
            "updated_at": project.updated_at.isoformat(),
            "master_data_count": 0
        }
    except Exception as e:
        logger.error(f"Error creating project: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error creating project: {str(e)}")

@app.get("/projects")
async def get_projects():
    """Get all projects"""
    try:
        db = SessionLocal()
        projects = db.query(Project).all()
        db.close()
        
        result = []
        for project in projects:
            # Get master data count
            db = SessionLocal()
            count = db.query(MasterDataItem).filter(MasterDataItem.project_id == project.id).count()
            db.close()
            
            result.append({
                "id": str(project.id),
                "name": project.name,
                "description": project.description,
                "created_at": project.created_at.isoformat(),
                "updated_at": project.updated_at.isoformat(),
                "master_data_count": count
            })
        
        return result
    except Exception as e:
        logger.error(f"Error fetching projects: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error fetching projects: {str(e)}")

@app.get("/projects/{project_id}")
async def get_project(project_id: str):
    """Get a specific project"""
    try:
        db = SessionLocal()
        project = db.query(Project).filter(Project.id == project_id).first()
        if not project:
            raise HTTPException(status_code=404, detail="Project not found")
        
        count = db.query(MasterDataItem).filter(MasterDataItem.project_id == project.id).count()
        db.close()
        
        return {
            "id": str(project.id),
            "name": project.name,
            "description": project.description,
            "created_at": project.created_at.isoformat(),
            "updated_at": project.updated_at.isoformat(),
            "master_data_count": count
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching project: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error fetching project: {str(e)}")

@app.post("/projects/{project_id}/master-data")
async def add_master_data(project_id: str, data: List[Dict[str, Any]]):
    """Add master data to a project"""
    try:
        db = SessionLocal()
        
        # Verify project exists
        project = db.query(Project).filter(Project.id == project_id).first()
        if not project:
            raise HTTPException(status_code=404, detail="Project not found")
        
        # Add master data items
        for item in data:
            master_item = MasterDataItem(
                project_id=project_id,
                item_description=item.get("item_description", ""),
                serial_number=item.get("serial_number", ""),
                tag_number=item.get("tag_number", ""),
                quantity=item.get("quantity", 1),
                status=item.get("status", "New")
            )
            db.add(master_item)
        
        db.commit()
        db.close()
        
        return {"message": f"Added {len(data)} items to master data"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error adding master data: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error adding master data: {str(e)}")

@app.get("/projects/{project_id}/master-data")
async def get_master_data(project_id: str):
    """Get master data for a project"""
    try:
        db = SessionLocal()
        items = db.query(MasterDataItem).filter(MasterDataItem.project_id == project_id).all()
        db.close()
        
        result = []
        for item in items:
            result.append({
                "id": str(item.id),
                "item_description": item.item_description,
                "serial_number": item.serial_number,
                "tag_number": item.tag_number,
                "quantity": item.quantity,
                "status": item.status,
                "created_at": item.created_at.isoformat(),
                "updated_at": item.updated_at.isoformat()
            })
        
        return result
    except Exception as e:
        logger.error(f"Error fetching master data: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error fetching master data: {str(e)}")

@app.post("/projects/{project_id}/match-data")
async def find_matches(project_id: str, request: Dict[str, Any]):
    """Find matches between extracted data and master data using enhanced SERIAL_NUMBER and Tag_Number matching"""
    try:
        extractedData = request.get("extractedData", [])
        datasetData = request.get("datasetData", [])
        match_strategy = request.get("matchStrategy", "serial_and_tag")
        confidence_threshold = request.get("confidenceThreshold", 0.5)
        
        db = SessionLocal()
        master_items = db.query(MasterDataItem).filter(MasterDataItem.project_id == project_id).all()
        db.close()
        
        # Combine master data with dataset data
        all_reference_items = []
        
        # Add master data
        for item in master_items:
            all_reference_items.append({
                "id": f"master_{item.id}",
                "item_description": item.item_description,
                "serial_number": item.serial_number,
                "tag_number": item.tag_number,
                "quantity": item.quantity,
                "status": item.status,
                "source": "master_data"
            })
        
        # Add dataset data
        for item in datasetData:
            if isinstance(item, dict):
                all_reference_items.append({
                    "id": f"dataset_{hash(str(item))}",
                    "item_description": item.get("description") or item.get("Description") or item.get("item_description", ""),
                    "serial_number": item.get("SERIAL_NUMBER") or item.get("serial_number", ""),
                    "tag_number": item.get("Tag_Number") or item.get("tag_number", ""),
                    "quantity": item.get("quantity") or item.get("Quantity", 1),
                    "status": item.get("status") or item.get("Status", "active"),
                    "source": "dataset"
                })
        
        matches = []
        for extracted_item in extractedData:
            item_matches = []
            serial_number = extracted_item.get("Serial_Number", "").strip().upper()
            tag_number = extracted_item.get("Tag_Number", "").strip().upper()
            
            for ref_item in all_reference_items:
                master_serial = ref_item["serial_number"].strip().upper()
                master_tag = ref_item["tag_number"].strip().upper()
                
                confidence = 0
                match_type = "none"
                
                # Enhanced matching logic for SERIAL_NUMBER and Tag_Number
                if serial_number and tag_number and master_serial and master_tag:
                    # Both exact match (highest priority)
                    if serial_number == master_serial and tag_number == master_tag:
                        confidence = 100
                        match_type = "both_exact"
                    # Serial exact, tag partial
                    elif serial_number == master_serial and (tag_number in master_tag or master_tag in tag_number):
                        confidence = 90
                        match_type = "serial_exact_tag_partial"
                    # Tag exact, serial partial
                    elif tag_number == master_tag and (serial_number in master_serial or master_serial in serial_number):
                        confidence = 85
                        match_type = "tag_exact_serial_partial"
                    # Both partial
                    elif (serial_number in master_serial or master_serial in serial_number) and (tag_number in master_tag or master_tag in tag_number):
                        confidence = 75
                        match_type = "both_partial"
                    # Only serial exact
                    elif serial_number == master_serial:
                        confidence = 80
                        match_type = "serial_only"
                    # Only tag exact
                    elif tag_number == master_tag:
                        confidence = 70
                        match_type = "tag_only"
                    # Only serial partial
                    elif serial_number in master_serial or master_serial in serial_number:
                        confidence = 60
                        match_type = "serial_partial"
                    # Only tag partial
                    elif tag_number in master_tag or master_tag in tag_number:
                        confidence = 50
                        match_type = "tag_partial"
                
                # Apply confidence threshold
                if confidence >= (confidence_threshold * 100):
                    item_matches.append({
                        "item": {
                            "id": ref_item["id"],
                            "item_description": ref_item["item_description"],
                            "serial_number": ref_item["serial_number"],
                            "tag_number": ref_item["tag_number"],
                            "quantity": ref_item["quantity"],
                            "status": ref_item["status"],
                            "source": ref_item["source"]
                        },
                        "confidence": confidence,
                        "match_type": match_type,
                        "extracted_serial": serial_number,
                        "extracted_tag": tag_number,
                        "master_serial": master_serial,
                        "master_tag": master_tag
                    })
            
            # Sort by confidence and match type priority
            item_matches.sort(key=lambda x: (
                x["confidence"],
                x["match_type"] in ["both_exact", "serial_exact_tag_partial", "tag_exact_serial_partial"]
            ), reverse=True)
            
            matches.append(item_matches)
        
        return matches
    except Exception as e:
        logger.error(f"Error finding matches: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error finding matches: {str(e)}")

@app.post("/projects/{project_id}/upload-excel")
async def upload_excel_dataset(project_id: str, file: UploadFile = File(...)):
    """Upload Excel file as master dataset"""
    try:
        # Verify project exists
        db = SessionLocal()
        project = db.query(Project).filter(Project.id == project_id).first()
        if not project:
            raise HTTPException(status_code=404, detail="Project not found")
        
        # For now, return a simple response since pandas is not available
        # In production, you would implement Excel parsing here
        db.close()
        
        return {"message": "Excel upload temporarily disabled - pandas not available. Please add data manually through the API."}
        
    except Exception as e:
        logger.error(f"Error uploading Excel: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error uploading Excel: {str(e)}")

@app.get("/projects/{project_id}/export-excel")
async def export_to_excel(project_id: str, format: str = "screen_cpu"):
    """Export project data to Excel in the specified format"""
    try:
        db = SessionLocal()
        items = db.query(MasterDataItem).filter(MasterDataItem.project_id == project_id).all()
        db.close()
        
        # Create Excel file using openpyxl directly
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        import io
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        
        if format == "screen_cpu":
            # Create the exact format from your image
            screen_items = [item for item in items if 'SCREEN' in item.item_description.upper() or 'MONITOR' in item.item_description.upper()]
            cpu_items = [item for item in items if 'CPU' in item.item_description.upper() or 'COMPUTER' in item.item_description.upper()]
            
            # Headers
            headers = ["Screen", "serial number", "Codification", "CPU", "Serial number", "codification"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Data rows
            max_rows = max(len(screen_items), len(cpu_items))
            for row in range(max_rows):
                row_num = row + 2
                
                # Screen data
                if row < len(screen_items):
                    ws.cell(row=row_num, column=1, value="SCREEN")
                    ws.cell(row=row_num, column=2, value=screen_items[row].serial_number)
                    ws.cell(row=row_num, column=3, value=screen_items[row].tag_number)
                
                # CPU data
                if row < len(cpu_items):
                    ws.cell(row=row_num, column=4, value="CPU")
                    ws.cell(row=row_num, column=5, value=cpu_items[row].serial_number)
                    ws.cell(row=row_num, column=6, value=cpu_items[row].tag_number)
            
        else:
            # Standard format
            headers = ["Item_Description", "Serial_Number", "Tag_Number", "Quantity", "Status"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            for row, item in enumerate(items, 2):
                ws.cell(row=row, column=1, value=item.item_description)
                ws.cell(row=row, column=2, value=item.serial_number)
                ws.cell(row=row, column=3, value=item.tag_number)
                ws.cell(row=row, column=4, value=item.quantity)
                ws.cell(row=row, column=5, value=item.status)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            io.BytesIO(output.read()),
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={"Content-Disposition": f"attachment; filename=project_{project_id}_export.xlsx"}
        )
            
    except Exception as e:
        logger.error(f"Error exporting Excel: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error exporting Excel: {str(e)}")

# Dataset endpoints
@app.post("/datasets")
async def create_dataset(request: Dict[str, Any]):
    """Create a new dataset"""
    db = None
    try:
        name = request.get("name")
        description = request.get("description", "")
        files = request.get("files", [])
        data = request.get("data", [])
        
        if not name:
            raise HTTPException(status_code=400, detail="Dataset name is required")
        
        if not isinstance(files, list):
            files = []
        if not isinstance(data, list):
            data = []
        
        logger.info(f"Creating dataset: {name} with {len(files)} files and {len(data)} rows")
        
        db = SessionLocal()
        dataset = Dataset(
            name=name,
            description=description,
            file_count=len(files),
            total_rows=len(data),
            files=json.dumps(files),
            data=json.dumps(data)
        )
        db.add(dataset)
        db.commit()
        db.refresh(dataset)
        
        result = {
            "id": str(dataset.id),
            "name": dataset.name,
            "description": dataset.description,
            "fileCount": dataset.file_count,
            "totalRows": dataset.total_rows,
            "files": files,
            "created_at": dataset.created_at.isoformat(),
            "updated_at": dataset.updated_at.isoformat()
        }
        
        logger.info(f"Dataset created successfully: {result['id']}")
        return result
    except Exception as e:
        logger.error(f"Error creating dataset: {str(e)}")
        if db:
            db.rollback()
        raise HTTPException(status_code=500, detail=f"Error creating dataset: {str(e)}")
    finally:
        if db:
            db.close()

@app.get("/datasets")
async def get_datasets():
    """Get all datasets"""
    db = None
    try:
        db = SessionLocal()
        datasets = db.query(Dataset).all()
        
        result = []
        for dataset in datasets:
            try:
                files = json.loads(dataset.files) if dataset.files else []
                result.append({
                    "id": str(dataset.id),
                    "name": dataset.name,
                    "description": dataset.description,
                    "fileCount": dataset.file_count,
                    "totalRows": dataset.total_rows,
                    "files": files,
                    "created_at": dataset.created_at.isoformat(),
                    "updated_at": dataset.updated_at.isoformat()
                })
            except Exception as e:
                logger.error(f"Error processing dataset {dataset.id}: {str(e)}")
                # Add dataset with empty files if JSON parsing fails
                result.append({
                    "id": str(dataset.id),
                    "name": dataset.name,
                    "description": dataset.description,
                    "fileCount": dataset.file_count,
                    "totalRows": dataset.total_rows,
                    "files": [],
                    "created_at": dataset.created_at.isoformat(),
                    "updated_at": dataset.updated_at.isoformat()
                })
        
        logger.info(f"Retrieved {len(result)} datasets")
        return result
    except Exception as e:
        logger.error(f"Error fetching datasets: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error fetching datasets: {str(e)}")
    finally:
        if db:
            db.close()

@app.get("/datasets/{dataset_id}")
async def get_dataset(dataset_id: str):
    """Get a specific dataset with its data"""
    try:
        db = SessionLocal()
        dataset = db.query(Dataset).filter(Dataset.id == dataset_id).first()
        if not dataset:
            raise HTTPException(status_code=404, detail="Dataset not found")
        
        files = json.loads(dataset.files) if dataset.files else []
        data = json.loads(dataset.data) if dataset.data else []
        db.close()
        
        return {
            "id": str(dataset.id),
            "name": dataset.name,
            "description": dataset.description,
            "fileCount": dataset.file_count,
            "totalRows": dataset.total_rows,
            "files": files,
            "data": data,
            "created_at": dataset.created_at.isoformat(),
            "updated_at": dataset.updated_at.isoformat()
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching dataset: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error fetching dataset: {str(e)}")

@app.post("/datasets/{dataset_id}/upload-files")
async def upload_dataset_files(dataset_id: str, files: List[UploadFile] = File(...)):
    """Upload files to an existing dataset"""
    try:
        db = SessionLocal()
        dataset = db.query(Dataset).filter(Dataset.id == dataset_id).first()
        if not dataset:
            raise HTTPException(status_code=404, detail="Dataset not found")
        
        # Process uploaded files
        combined_data = []
        existing_files = json.loads(dataset.files) if dataset.files else []
        new_files = []
        
        for file in files:
            try:
                content = await file.read()
                
                # Parse Excel/CSV files properly
                if file.filename.endswith('.csv'):
                    df = pd.read_csv(io.BytesIO(content))
                elif file.filename.endswith(('.xlsx', '.xls')):
                    df = pd.read_excel(io.BytesIO(content))
                else:
                    logger.warning(f"Unsupported file type: {file.filename}")
                    continue
                
                # Convert to list of dictionaries
                file_data = df.to_dict('records')
                combined_data.extend(file_data)
                new_files.append(file.filename)
                
                logger.info(f"Processed {file.filename}: {len(file_data)} rows")
                
            except Exception as e:
                logger.error(f"Error processing file {file.filename}: {str(e)}")
        
        # Update dataset
        all_files = existing_files + new_files
        existing_data = json.loads(dataset.data) if dataset.data else []
        all_data = existing_data + combined_data
        
        dataset.file_count = len(all_files)
        dataset.total_rows = len(all_data)
        dataset.files = json.dumps(all_files)
        dataset.data = json.dumps(all_data)
        dataset.updated_at = datetime.utcnow()
        
        db.commit()
        db.close()
        
        return {
            "message": f"Uploaded {len(new_files)} files to dataset",
            "totalFiles": len(all_files),
            "totalRows": len(all_data)
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error uploading files to dataset: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error uploading files to dataset: {str(e)}")

@app.delete("/datasets/{dataset_id}")
async def delete_dataset(dataset_id: str):
    """Delete a dataset"""
    try:
        db = SessionLocal()
        dataset = db.query(Dataset).filter(Dataset.id == dataset_id).first()
        if not dataset:
            raise HTTPException(status_code=404, detail="Dataset not found")
        
        db.delete(dataset)
        db.commit()
        db.close()
        
        return {"message": "Dataset deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting dataset: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error deleting dataset: {str(e)}")

@app.post("/extract-image")
async def extract_image_data(file: UploadFile = File(...)):
    """Extract data from image and format as table"""
    try:
        if not file.content_type.startswith('image/'):
            raise HTTPException(status_code=400, detail="File must be an image")
        
        # Read image data
        image_data = await file.read()
        
        # Get API key
        api_key = get_next_api_key()
        
        # Extract text
        extracted_text = await extract_text_from_image(image_data, api_key)
        
        # Format as table
        table_data = await format_text_as_table(extracted_text, api_key)
        
        return {
            "success": True,
            "extracted_text": extracted_text,
            "table_data": table_data,
            "filename": file.filename,
            "item_count": len(table_data)
        }
        
    except Exception as e:
        logger.error(f"Error extracting image data: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error extracting image data: {str(e)}")

@app.post("/search")
async def search_data(request: Dict[str, Any]):
    """Search across all datasets and master data"""
    try:
        query = request.get("query", "").strip()
        item_type = request.get("item_type", "").strip()
        status = request.get("status", "").strip()
        source = request.get("source", "").strip()
        date_from = request.get("date_from", "").strip()
        date_to = request.get("date_to", "").strip()
        min_quantity = request.get("min_quantity")
        max_quantity = request.get("max_quantity")
        exact_match = request.get("exact_match", True)
        case_sensitive = request.get("case_sensitive", False)
        page = request.get("page", 1)
        limit = request.get("limit", 20)
        
        db = SessionLocal()
        # Aggregate results for pagination
        all_results: List[Dict[str, Any]] = []
        
        # Search master data
        if not source or source == "master_data":
            master_query = db.query(MasterDataItem)
            
            # Apply filters
            if query:
                if exact_match:
                    if case_sensitive:
                        master_query = master_query.filter(
                            (MasterDataItem.item_description == query) |
                            (MasterDataItem.serial_number == query) |
                            (MasterDataItem.tag_number == query)
                        )
                    else:
                        master_query = master_query.filter(
                            (MasterDataItem.item_description.ilike(query)) |
                            (MasterDataItem.serial_number.ilike(query)) |
                            (MasterDataItem.tag_number.ilike(query))
                        )
                else:
                    if case_sensitive:
                        master_query = master_query.filter(
                            (MasterDataItem.item_description.contains(query)) |
                            (MasterDataItem.serial_number.contains(query)) |
                            (MasterDataItem.tag_number.contains(query))
                        )
                    else:
                        master_query = master_query.filter(
                            (MasterDataItem.item_description.ilike(f"%{query}%")) |
                            (MasterDataItem.serial_number.ilike(f"%{query}%")) |
                            (MasterDataItem.tag_number.ilike(f"%{query}%"))
                        )
            
            if status:
                master_query = master_query.filter(MasterDataItem.status.ilike(f"%{status}%"))
            
            if min_quantity is not None:
                master_query = master_query.filter(MasterDataItem.quantity >= min_quantity)
            
            if max_quantity is not None:
                master_query = master_query.filter(MasterDataItem.quantity <= max_quantity)
            
            master_items = master_query.all()
            
            for item in master_items:
                # Check item type filter
                if item_type:
                    item_desc_lower = item.item_description.lower()
                    if item_type == "cpu" and not ("cpu" in item_desc_lower or "computer" in item_desc_lower):
                        continue
                    elif item_type == "screen" and not ("screen" in item_desc_lower or "monitor" in item_desc_lower or "display" in item_desc_lower):
                        continue
                    elif item_type == "other" and ("cpu" in item_desc_lower or "computer" in item_desc_lower or "screen" in item_desc_lower or "monitor" in item_desc_lower or "display" in item_desc_lower):
                        continue
                
                # Get project name
                project = db.query(Project).filter(Project.id == item.project_id).first()
                project_name = project.name if project else "Unknown Project"
                
                all_results.append({
                    "id": f"master_{item.id}",
                    "item_description": item.item_description,
                    "serial_number": item.serial_number,
                    "tag_number": item.tag_number,
                    "quantity": item.quantity,
                    "status": item.status,
                    "source": "master_data",
                    "project_name": project_name,
                    "created_at": item.created_at.isoformat()
                })
        
        # Search datasets
        if not source or source == "dataset":
            datasets = db.query(Dataset).all()
            
            for dataset in datasets:
                try:
                    dataset_data = json.loads(dataset.data) if dataset.data else []
                    
                    for item in dataset_data:
                        if not isinstance(item, dict):
                            continue
                        
                        # Apply filters - handle different column name formats
                        item_desc = (item.get("Item_Description") or 
                                   item.get("item_description") or 
                                   item.get("Description") or 
                                   item.get("ITEM") or
                                   item.get("Unnamed: 1", ""))
                        serial_num = (item.get("Serial_Number") or 
                                    item.get("serial_number") or 
                                    item.get("SERIAL_NUMBER") or
                                    item.get("Unnamed: 3", ""))
                        tag_num = (item.get("Tag_Number") or 
                                 item.get("tag_number") or 
                                 item.get("Tag_Number") or
                                 item.get("CODE") or
                                 item.get("CODIFICATION FOR MOH") or
                                 item.get("Unnamed: 2", ""))
                        item_status = item.get("Status") or item.get("status", "active")
                        quantity = item.get("Quantity") or item.get("quantity", 1)
                        
                        # Check query filter
                        if query:
                            search_text = f"{item_desc} {serial_num} {tag_num}".lower() if not case_sensitive else f"{item_desc} {serial_num} {tag_num}"
                            query_text = query.lower() if not case_sensitive else query
                            
                            if exact_match:
                                if not (item_desc == query or serial_num == query or tag_num == query):
                                    continue
                            else:
                                if query_text not in search_text:
                                    continue
                        
                        # Check item type filter
                        if item_type:
                            item_desc_lower = item_desc.lower()
                            if item_type == "cpu" and not ("cpu" in item_desc_lower or "computer" in item_desc_lower):
                                continue
                            elif item_type == "screen" and not ("screen" in item_desc_lower or "monitor" in item_desc_lower or "display" in item_desc_lower):
                                continue
                            elif item_type == "other" and ("cpu" in item_desc_lower or "computer" in item_desc_lower or "screen" in item_desc_lower or "monitor" in item_desc_lower or "display" in item_desc_lower):
                                continue
                        
                        # Check status filter
                        if status and status.lower() not in item_status.lower():
                            continue
                        
                        # Check quantity filters
                        try:
                            qty = int(quantity) if quantity else 1
                            if min_quantity is not None and qty < min_quantity:
                                continue
                            if max_quantity is not None and qty > max_quantity:
                                continue
                        except (ValueError, TypeError):
                            qty = 1
                        
                        all_results.append({
                            "id": f"dataset_{dataset.id}_{hash(str(item))}",
                            "item_description": item_desc,
                            "serial_number": serial_num,
                            "tag_number": tag_num,
                            "quantity": qty,
                            "status": item_status,
                            "source": "dataset",
                            "dataset_name": dataset.name,
                            "created_at": dataset.created_at.isoformat()
                        })
                        
                except Exception as e:
                    logger.error(f"Error processing dataset {dataset.id}: {str(e)}")
                    continue
        
        db.close()
        
        # Calculate pagination
        total_results = len(all_results)
        start_idx = (page - 1) * limit
        end_idx = start_idx + limit
        paginated_results = all_results[start_idx:end_idx]
        total_pages = (total_results + limit - 1) // limit
        
        return {
            "success": True,
            "results": paginated_results,
            "total_results": total_results,
            "total_pages": total_pages,
            "current_page": page,
            "limit": limit
        }
        
    except Exception as e:
        logger.error(f"Error searching data: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error searching data: {str(e)}")

@app.post("/bulk-search")
async def bulk_search(request: Dict[str, Any]):
    """Bulk search across all datasets and master data"""
    try:
        search_terms = request.get("search_terms", [])
        item_type = request.get("item_type", "").strip()
        status = request.get("status", "").strip()
        source = request.get("source", "").strip()
        exact_match = request.get("exact_match", True)
        case_sensitive = request.get("case_sensitive", False)
        restrict_to = request.get("restrict_to", "serial_or_code")  # serial_or_code | all
        
        if not search_terms:
            return {"success": True, "results": [], "total_results": 0, "total_pages": 0, "current_page": 1, "limit": 0}
        
        def _norm(v):
            return v if case_sensitive else str(v or "").lower()
        
        normalized_terms = [_norm(str(t).strip()) for t in search_terms if str(t).strip()]
        # Build exact lookup maps for codification (tag) and serial
        def _prefer_dataset(existing: Optional[Dict[str, Any]], candidate: Dict[str, Any]) -> Dict[str, Any]:
            if existing is None:
                return candidate
            if existing.get("source") == "dataset" and candidate.get("source") != "dataset":
                return existing
            if existing.get("source") != "dataset" and candidate.get("source") == "dataset":
                return candidate
            return existing

        code_map: Dict[str, Dict[str, Any]] = {}
        serial_map: Dict[str, Dict[str, Any]] = {}

        db = SessionLocal()
        try:
            # Index datasets
            if not source or source == "dataset":
                for dataset in db.query(Dataset).all():
                    dataset_data = json.loads(dataset.data) if dataset.data else []
                    for item in dataset_data:
                        if not isinstance(item, dict):
                            continue
                        item_desc = (item.get("Item_Description") or item.get("item_description") or item.get("Description") or item.get("ITEM") or item.get("Unnamed: 1", ""))
                        serial_num = (item.get("Serial_Number") or item.get("serial_number") or item.get("SERIAL_NUMBER") or item.get("Unnamed: 3", ""))
                        tag_num = (item.get("Tag_Number") or item.get("tag_number") or item.get("Tag_Number") or item.get("CODE") or item.get("CODIFICATION FOR MOH") or item.get("Unnamed: 2", ""))
                        item_status = item.get("Status") or item.get("status", "active")
                        quantity = item.get("Quantity") or item.get("quantity", 1)

                        # Apply filters (exact-only bulk search)
                        if item_type and item_type.lower() not in (item_desc or "").lower():
                            continue
                        if status and status.lower() != (item_status or "").lower():
                            continue

                        n_tag = _norm(tag_num)
                        n_serial = _norm(serial_num)
                        base_result = {
                            "id": f"dataset_{dataset.id}_{hash(str(item))}",
                            "item_description": item_desc,
                            "serial_number": serial_num,
                            "tag_number": tag_num,
                            "quantity": quantity,
                            "status": item_status,
                            "source": "dataset",
                            "dataset_name": dataset.name,
                            "created_at": dataset.created_at.isoformat()
                        }
                        if n_tag:
                            code_map[n_tag] = _prefer_dataset(code_map.get(n_tag), base_result)
                        if n_serial:
                            serial_map[n_serial] = _prefer_dataset(serial_map.get(n_serial), base_result)

            # Index master data
            if not source or source == "master_data":
                for item in db.query(MasterDataItem).all():
                    # Apply filters
                    if item_type and item_type.lower() not in (item.item_description or "").lower():
                        continue
                    if status and status.lower() != (item.status or "").lower():
                        continue

                    n_tag = _norm(item.tag_number)
                    n_serial = _norm(item.serial_number)
                    base_result = {
                        "id": f"master_{item.id}",
                        "item_description": item.item_description,
                        "serial_number": item.serial_number,
                        "tag_number": item.tag_number,
                        "quantity": item.quantity,
                        "status": item.status,
                        "source": "master_data",
                        "project_name": item.project.name if item.project else None,
                        "created_at": item.created_at.isoformat()
                    }
                    if n_tag:
                        code_map[n_tag] = _prefer_dataset(code_map.get(n_tag), base_result)
                    if n_serial:
                        serial_map[n_serial] = _prefer_dataset(serial_map.get(n_serial), base_result)
        finally:
            db.close()

        # Resolve strictly one result per input term by exact equality on code first, then serial
        now_iso = datetime.utcnow().isoformat()
        final_results: List[Dict[str, Any]] = []
        for idx, term in enumerate(normalized_terms):
            hit = code_map.get(term) or serial_map.get(term)
            if hit is not None:
                # Ensure stable unique id per input position
                hit = dict(hit)
                hit["id"] = f"{hit['source']}_{idx}_{abs(hash(term))}"
                final_results.append(hit)
            else:
                term_value = search_terms[idx]
                final_results.append({
                    "id": f"notfound_{idx}_{abs(hash(term_value))}",
                    "item_description": "Not Found",
                    "serial_number": "",
                    "tag_number": str(term_value),
                    "quantity": 0,
                    "status": "NOT FOUND",
                    "source": "none",
                    "created_at": now_iso
                })

        return {
            "success": True,
            "results": final_results,
            "total_results": len(final_results),
            "total_pages": 1,
            "current_page": 1,
            "limit": len(final_results)
        }
    except Exception as e:
        logger.error(f"Error in bulk search: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error in bulk search: {str(e)}")

@app.post("/export-search-results")
async def export_search_results(request: Dict[str, Any]):
    """Export search results to Excel or CSV"""
    try:
        results = request.get("results", [])
        format_type = request.get("format", "excel")
        filename = request.get("filename", "search_results")
        
        if not results:
            raise HTTPException(status_code=400, detail="No results to export")
        
        # Convert results to DataFrame
        df = pd.DataFrame(results)
        
        if format_type == "excel":
            # Create Excel file using openpyxl
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            import io
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Search Results"
            
            # Headers
            headers = ["Item Description", "Serial Number", "Tag Number", "Quantity", "Status", "Source", "Dataset/Project", "Created At"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Data rows
            for row, result in enumerate(results, 2):
                ws.cell(row=row, column=1, value=result.get("item_description", ""))
                ws.cell(row=row, column=2, value=result.get("serial_number", ""))
                ws.cell(row=row, column=3, value=result.get("tag_number", ""))
                ws.cell(row=row, column=4, value=result.get("quantity", ""))
                ws.cell(row=row, column=5, value=result.get("status", ""))
                ws.cell(row=row, column=6, value=result.get("source", ""))
                ws.cell(row=row, column=7, value=result.get("dataset_name", "") or result.get("project_name", ""))
                ws.cell(row=row, column=8, value=result.get("created_at", ""))
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to BytesIO
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            return StreamingResponse(
                io.BytesIO(output.read()),
                media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"}
            )
        
        else:  # CSV
            # Create CSV content
            csv_content = df.to_csv(index=False)
            
            # Create temporary file
            with tempfile.NamedTemporaryFile(mode='w', delete=False, suffix='.csv', encoding='utf-8') as tmp_file:
                tmp_file.write(csv_content)
                
                return FileResponse(
                    tmp_file.name,
                    media_type='text/csv',
                    filename=f"{filename}.csv"
                )
                
    except Exception as e:
        logger.error(f"Error exporting search results: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error exporting search results: {str(e)}")

@app.delete("/datasets/{dataset_id}")
async def delete_dataset(dataset_id: str):
    """Delete a dataset"""
    try:
        db = SessionLocal()
        dataset = db.query(Dataset).filter(Dataset.id == dataset_id).first()
        if not dataset:
            raise HTTPException(status_code=404, detail="Dataset not found")
        
        db.delete(dataset)
        db.commit()
        db.close()
        
        return {"message": "Dataset deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting dataset: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error deleting dataset: {str(e)}")

@app.get("/health")
async def health_check():
    """Health check endpoint"""
    return {"status": "healthy", "timestamp": datetime.now().isoformat()}

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
